# Xuất báo cáo: Excel (.xlsx qua openpyxl, hoặc .csv), PDF (fpdf2 + font Unicode).
import csv
import os
import re
from pathlib import Path


def _unicode_ttf_path():
    try:
        import fpdf as fpdf_pkg

        bundled = Path(fpdf_pkg.__file__).resolve().parent / "font" / "DejaVuSans.ttf"
        if bundled.is_file():
            return bundled
    except ImportError:
        pass
    windir = Path(os.environ.get("WINDIR", r"C:\Windows"))
    for fn in ("arial.ttf", "Arial.ttf", "tahoma.ttf", "Tahoma.ttf", "segoeui.ttf"):
        cand = windir / "Fonts" / fn
        if cand.is_file():
            return cand
    return None


def _sheet_name_safe(name: str) -> str:
    s = re.sub(r'[\[\]:\\/*?]', "_", str(name).strip()) or "BaoCao"
    return s[:31]


def export_to_excel(path: str | Path, sheet_title: str, headers: list, rows: list) -> tuple[str, str | None]:
    """
    Trả về (đuôi file thực tế 'xlsx' hoặc 'csv', thông báo lỗi nếu có).
    Nếu không có openpyxl: ghi CSV UTF-8 BOM (Excel mở được), đổi đuôi .xlsx -> .csv.
    """
    path = Path(path)
    headers = [str(h) for h in headers]
    norm_rows = [[_cell_export(v) for v in r] for r in rows]

    try:
        from openpyxl import Workbook

        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = _sheet_name_safe(sheet_title)
        ws.append(headers)
        for r in norm_rows:
            ws.append(r)
        wb.save(str(path))
        return "xlsx", None
    except ImportError:
        out = path.with_suffix(".csv") if path.suffix.lower() == ".xlsx" else path
        try:
            with open(out, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(headers)
                w.writerows(norm_rows)
            return "csv", None
        except OSError as e:
            return "", str(e)
    except OSError as e:
        return "", str(e)


def _cell_export(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return v
    return str(v)


def export_to_pdf(path: str | Path, title: str, headers: list, rows: list) -> str | None:
    """None nếu thành công; chuỗi lỗi nếu thất bại."""
    try:
        from fpdf import FPDF
    except ImportError:
        return "Chưa cài fpdf2. Chạy: pip install fpdf2"

    font_file = _unicode_ttf_path()
    if not font_file:
        return "Không tìm thấy font Unicode (Arial/Tahoma trên Windows hoặc fpdf/font)."

    path = Path(path)
    if path.suffix.lower() != ".pdf":
        path = path.with_suffix(".pdf")

    headers = [str(h) for h in headers]
    norm_rows = [[_cell_export(v) for v in r] for r in rows]

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.add_font("DejaVu", "", str(font_file))
    pdf.set_font("DejaVu", size=11)
    pdf.multi_cell(0, 8, str(title), align="C")
    pdf.ln(2)

    n = max(len(headers), 1)
    usable = pdf.w - 2 * pdf.l_margin
    col_w = usable / n
    line_h = 7
    pdf.set_font("DejaVu", size=8)
    for h in headers:
        pdf.cell(col_w, line_h, str(h)[:48], border=1, align="C")
    pdf.ln()
    pdf.set_font("DejaVu", size=7)
    for r in norm_rows:
        if pdf.get_y() > pdf.h - 20:
            pdf.add_page()
            pdf.set_font("DejaVu", size=8)
            for h in headers:
                pdf.cell(col_w, line_h, str(h)[:48], border=1, align="C")
            pdf.ln()
            pdf.set_font("DejaVu", size=7)
        for cell in r:
            t = str(cell).replace("\n", " ")
            if len(t) > 42:
                t = t[:39] + "..."
            pdf.cell(col_w, line_h, t, border=1)
        pdf.ln()

    try:
        pdf.output(str(path))
    except OSError as e:
        return str(e)
    return None
